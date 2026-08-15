#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""金融创新全球作战地图 —— 生成器流水线（单一离线 HTML）。
用法：
  .venv/Scripts/python build.py init                 # 由示例生成 news.xlsx（可编辑数据源，标记示例）
  .venv/Scripts/python build.py build                # 读 news.xlsx -> 生成单文件 HTML（数据内联快照）
  .venv/Scripts/python build.py build --source supabase   # 先从 Supabase 拉取再生成
  .venv/Scripts/python build.py build --source sample     # 用内置示例数据生成（标记示例）
  .venv/Scripts/python build.py pull                 # Supabase -> news.xlsx
  .venv/Scripts/python build.py push                # news.xlsx -> Supabase（按 title upsert，自动去重）
  .venv/Scripts/python build.py merge <json文件>     # 把真实新闻并入 news.xlsx，已存在的自动跳过（去重）
  .venv/Scripts/python build.py replace <json文件>   # 只移除示例行，用真实新闻替换（保留已有真实数据，按内容去重）
"""
import argparse, json, pathlib, re, sys
from collections import Counter
import openpyxl

HERE = pathlib.Path(__file__).resolve().parent
# 只读技能资产（随技能分发，安装后不应被改写）
TEMPLATE = HERE / "template.html"
ASSETS = HERE / "assets"                 # echarts.min.js / world.json / china.json
SAMPLE = HERE / "news_sample.json"       # 内置示例数据源
# 数据与产物落在「用户当前工作目录」，避免写入技能安装目录（可分发、不污染自身）
WORKDIR = pathlib.Path.cwd()
XLSX = WORKDIR / "news.xlsx"
OUT_HTML = WORKDIR / "金融创新全球作战地图.html"

COLUMNS = ["time","category","direction","company","city","lng","lat","title","link","event","why","innovation","learn","is_sample"]

# ---------------------------------------------------------------------------
# 建议区（右侧面板）内容生成 —— 数据驱动；每次 build 都基于当前新闻重新计算
# 趋势与建议为针对 20 个子领域的"实战洞察"，非通用模板话术。
# ---------------------------------------------------------------------------
CAT_COLOR = {"银行":"#26e0ff","保险":"#b388ff","金融科技":"#ffb43d"}

# 每个子方向（direction）对应的「趋势判断 + 高管动作建议」
DIRECTION_INSIGHTS = {
    "信用卡分期": {
        "trend": "信用卡分期从大额消费分期转向「场景化小额高频 + 会员权益捆绑」，银行以真实消费场景提升活卡率与中间业务收入，告别单纯价格战。",
        "advice": "把分期嵌入出行/教育/家装等真实场景，用权益而非利率做差异化；考核活卡率与中收，而非只看分期余额规模。"},
    "普惠保险": {
        "trend": "惠民保/城市定制险从「跑马圈地」进入「可持续经营」，多地收紧免赔与续保规则，行业重心从获客转向留存与精算平衡。",
        "advice": "从获客转向留存与健康管理绑定；与医保局共建数据闭环控制赔付率；城市定制险须有财政与医保的可持续背书。"},
    "寿险": {
        "trend": "寿险从储蓄型主推转向「保障+养老+分红」组合，分红险与增额终身寿在利率下行期承接居民稳健配置需求。",
        "advice": "产品组合化，用分红险平滑利差损；代理人渠道聚焦中高净值客户养老传承需求，做长期价值而非短缴冲刺。"},
    "公募私募": {
        "trend": "银行代销从「卖产品」转向「投顾+配置」，银证联动与基金投顾试点扩容，买方投顾能力成为 AUM 竞争核心。",
        "advice": "建立买方投顾能力，用账户制做客户资产配置、沉淀 AUM 而非单笔佣金；把基金代销做成陪伴式服务。"},
    "存款理财": {
        "trend": "利率下行与手工补息整改后，存款理财转向「现金管理+固收+」与结构性存款，理财替代部分存款成为客户现金流管理工具。",
        "advice": "用理财替代部分存款、做客户现金流管理；结构性存款做收益增强但须合规披露，严禁变相高息揽储。"},
    "个人信用贷款": {
        "trend": "消费贷贴息与利率下行叠加，银行以「消费贷+场景」抢优质客群，但需警惕共债与资金挪用风险。",
        "advice": "用数据做客群分层与额度动态管理；贴息获客后靠交叉销售回收成本；建立共债监测与资金用途校验。"},
    "财险": {
        "trend": "财险创新集中于车险新能源化、农险科技化、责任险场景化，车电分离、生猪期货等新模式降低赔付波动。",
        "advice": "用 IoT 与遥感做风险减量；农险/责任险做场景嵌入降低赔付率；新能源车企合作须重构定价模型。"},
    "养老险": {
        "trend": "养老金融政策驱动，个人养老金、商业养老险、养老 Y 份额扩容，银行保险共建养老账户生态。",
        "advice": "把养老账户与银行养老理财/保险做账户联通；做长期资金留存，用养老场景绑定客户家庭全生命周期。"},
    "个人房贷": {
        "trend": "房贷利率继续下行、首套优惠延续，但总量见顶，银行从「增量扩张」转向「存量按揭客户综合经营」。",
        "advice": "对存量按揭客户做综合经营（消费贷/财富）；用转按揭/置换做客户挽留，提升单客价值而非新增规模。"},
    "信托": {
        "trend": "银行+信托财富联动深化，家族信托、服务信托成为私行高净值客户资产配置与传承的标配工具。",
        "advice": "用家族信托做私行客户资产隔离与传承；与服务信托绑定企业财富管理，提升高净值客户黏性。"},
    "技术创新": {
        "trend": "大模型/AI 在风控、客服、投研、研发场景从 PoC 走向生产，数据底座与 AI 中台成为落地前提。",
        "advice": "先建数据底座与 AI 中台，把 AI 从「工具」变「生产力」；重定义可计算的非结构化资产，优先高 ROI 场景。"},
    "汽融": {
        "trend": "汽车金融聚焦新能源，贴息+融资租赁+残值管理并行，银行与头部车企、经销商深度绑定做一站式购车金融。",
        "advice": "做「车+金融+保险」一站式；用残值模型控制风险；深度绑定头部新能源车企与渠道，提升渗透率。"},
    "健康险": {
        "trend": "健康险从重疾转向「医疗+带病体+惠民」，长期医疗险与免健告产品扩容，健康管理成为差异化抓手。",
        "advice": "做「保险+健康管理」闭环，用可穿戴与理赔数据做风险减量；带病体产品须精算可控、防逆选择。"},
    "产品创新": {
        "trend": "智能体、AI 银行卡、数字员工产品化，银行推出 AI 原生产品，重构客户与银行的交互界面。",
        "advice": "做 AI 原生产品而非「旧产品+AI 壳」；重定义客户与银行交互界面，把非结构化需求变成可计算服务。"},
    "权益升级": {
        "trend": "会员权益从积分转向「AI 权益/场景权益」，银行用权益做客群分层运营，提升活跃与黏性。",
        "advice": "用权益做客群分层运营；权益与真实场景（出行/内容/健康）绑定，避免积分通胀沦为成本中心。"},
    "小微企业主贷款": {
        "trend": "小微普惠从「冲量」转向「提质」，用税务/发票/供应链数据做信用贷，经营贷抵押化与信用化并行。",
        "advice": "用税务/发票/订单数据建模，做首贷户拓展与无还本续贷；严控不良，做小微经营贷而非变相房贷。"},
    "银保": {
        "trend": "银保从趸交理财转向期交保障+养老，银行渠道价值被重估，从「通道」向「客户经营伙伴」升级。",
        "advice": "银保做期交与保障型产品；与险企共建客户经营；银行从通道变客户经营伙伴，提升中收质量。"},
    "模式创新": {
        "trend": "开放银行、供应链金融、数据要素流通、可信数据空间等新模式，把银行能力 API 化嵌入产业场景。",
        "advice": "用开放银行做生态嵌入；把银行能力 API 化嵌入产业场景；以可信数据空间打通跨机构数据协作。"},
    "需求走向": {
        "trend": "客户需求从「产品」转向「解决方案/陪伴式服务」，年轻客群与银发客群两极分化明显。",
        "advice": "做客群分层经营；陪伴式投顾与适老化并行；用数据识别生命周期节点做主动服务。"},
    "个人房地产抵押贷款": {
        "trend": "经营贷抵押化，房产抵押经营贷服务小微，但须严防资金违规流入楼市，监管穿透式核查趋严。",
        "advice": "用抵押物+经营流水做闭环，严控资金用途；做小微经营贷而非变相房贷，强化受托支付与贷后监测。"},
}

# 建议视角库（种子场景，可编辑/可扩展——仅供默认展示；运行时用户可输入任意自定义视角，
# 由前端基于下方 DIRECTION_INSIGHTS 趋势库按角色关键词动态生成建议。本字典只是开箱即用的起点，
# 绝非写死的固定受众集；"分行行长"只是其中一个示例场景，可随意增删改或完全自定义）。
ROLE_ADVICE = {
    "银行高管": {
        "color": "#26e0ff", "icon": "🏦",
        "points": [
            "<b>AI 从 PoC 到生产：</b>先建数据底座与 AI 中台，重定义「可计算的非结构化资产」，优先风控/客服/投研等高 ROI 场景。",
            "<b>净息差压力下：</b>用「理财+财富+中收」替代存款依赖，以账户制经营客户 AUM，而非单笔产品。",
            "<b>信贷从增量转存量经营：</b>用场景分期、经营贷、按揭客户综合经营做客户纵深，提升单客价值。",
            "<b>合规底线：</b>经营贷防资金挪用、消费贷防共债、代销重适当性，监管穿透核查趋严。",
        ]},
    "保险高管": {
        "color": "#b388ff", "icon": "🛡",
        "points": [
            "<b>寿险/养老承接利率下行需求：</b>产品组合化+分红平滑利差损，聚焦中高净值客户养老传承。",
            "<b>普惠/惠民保从获客转留存：</b>与医保局共建数据闭环控制赔付率，做可持续经营。",
            "<b>健康险做「保险+健康管理」风险减量闭环；</b>带病体产品须精算可控、防逆选择。",
            "<b>银保渠道做期交保障型：</b>银行从通道变客户经营伙伴，提升中收质量。",
        ]},
    "金融科技负责人": {
        "color": "#ffb43d", "icon": "⚡",
        "points": [
            "<b>大模型落地优先高 ROI 场景</b>（客服/风控/投研/研发），建 AI 中台而非单点，数据治理是前提。",
            "<b>做 AI 原生产品而非「旧产品+AI 壳」</b>，重定义客户与银行交互界面。",
            "<b>用开放银行/可信数据空间把能力 API 化嵌入产业</b>，做生态而非功能。",
            "<b>隐私计算与数据安全是规模化前提</b>，先把数据底座与合规框架搭稳。",
        ]},
    "金融高管视角": {
        "color": "#26e0ff", "icon": "🏛",
        "points": [
            "<b>综合金融是唯一能对抗网点劣势的杠杆：</b>以「一个客户、一个账户、多种产品、一站式服务」重定义客户关系，把银行/保险/理财/信贷的客群与数据打通，用账户制经营 AUM 而非单笔产品。",
            "<b>净息差下行期，中收与财富替代存款依赖：</b>银保期交保障、理财/基金投顾、私行与信托做高净值纵深；以「客户生命周期」为轴线配置存贷保财，提升单客价值。",
            "<b>AI 从 PoC 到生产，重定义「可计算的非结构化资产」：</b>优先风控/客服/投研/营销高 ROI 场景，建 AI 中台而非单点；用可信数据空间/隐私计算把跨牌照数据合规变成可计算资产。",
            "<b>合规与风险是跨牌照底线：</b>经营贷防资金挪用、代销重适当性、惠民保控赔付率、消保与反洗钱穿透核查趋严——综金越深，跨条线适当性与数据安全越要前置闭环。",
        ]},
    "分行行长（省级分行）": {
        "color": "#5ee0a0", "icon": "📍",
        "points": [
            "<b>区域市场份额优先于总行统一打法：</b>紧盯本地同业（国有大行/股份/城农商）在存款、按揭、普惠的份额变化，用「属地特色产业+对公链条」打差异化，而非照搬全行模板。",
            "<b>支行 KPI 与一线队伍是落地关键：</b>把总行指标拆成可执行的支行任务包，用场景分期/经营贷/代发工资等抓手给网点具体弹药；队伍能力决定落地，培训与激励须贴网点节奏。",
            "<b>属地监管与风险在本地闭环：</b>地方金融监管局与央行分行的窗口指导、消保与反洗钱检查都在属地，经营贷资金用途、代销适当性、消保投诉是高频风险点，须本地化合规闭环。",
            "<b>总行战略本地化选点做深：</b>承接总行 AI/中收/养老方向，但在本省选 1–2 个高 ROI 场景（如本地政务数据+普惠、区域车企汽融）做透，用本地数据做客户分层，避免摊大饼。",
        ]},
}

DEFAULT_ROLE = "金融高管视角"

# 自定义角色关键词→方向映射：build 时给定任意角色名，按关键词匹配最相关业务方向，
# 从 DIRECTION_INSIGHTS 动态提取建议（非写死），并固化进 HTML（不在浏览器交互选择）。
ROLE_KEYWORDS = [
    (re.compile(r"分行|区域|支行|地市|省分行|一线|网点|属地|零售条线"),
     ["个人信用贷款","小微企业主贷款","个人房地产抵押贷款","个人房贷","信用卡分期","存款理财","银保","汽融"]),
    (re.compile(r"监管|合规|风险|消保|反洗钱|合规官"),
     ["个人信用贷款","个人房地产抵押贷款","银保","存款理财","健康险"]),
    (re.compile(r"科技|数字|AI|大模型|数据|智能|信息"),
     ["技术创新","产品创新","模式创新","权益升级"]),
    (re.compile(r"保险|保障|养老|健康|寿险|财险|险"),
     ["寿险","养老险","健康险","财险","普惠保险","银保"]),
    (re.compile(r"财富|理财|私人|资管|高净值|信托|私行"),
     ["公募私募","信托","存款理财","养老险"]),
    (re.compile(r"普惠|小微|三农|实体|经营|企业主"),
     ["小微企业主贷款","个人信用贷款","普惠保险"]),
    (re.compile(r"汽车|车"), ["汽融"]),
    (re.compile(r"房|按揭|地产"), ["个人房贷","个人房地产抵押贷款"]),
]

def role_advice_for(role, fallback_dirs=None):
    """给定任意角色名，生成建议块（动态，非写死）；无匹配关键词时回退到当前热点趋势方向。"""
    picks = []
    for rx, dirs in ROLE_KEYWORDS:
        if rx.search(role):
            picks.extend(dirs)
    picks = list(dict.fromkeys(picks))
    if not picks:
        picks = list(fallback_dirs or [])[:4] or list(DIRECTION_INSIGHTS.keys())[:4]
    picks = picks[:4]
    points = [f"<b>{d}：</b>{DIRECTION_INSIGHTS.get(d, {}).get('advice', '')}" for d in picks]
    return {"audience": role, "color": "#5ee0a0", "icon": "🎯", "points": points}

def build_suggestions(news, role=None):
    """基于当前新闻计算建议区所需的结构化内容。role 由问题流程询问用户后在 build 时固化（--role）。"""
    if role is None:
        role = DEFAULT_ROLE
    total = len(news)
    by_cat = Counter(n.get("category") for n in news if n.get("category"))
    dir_counter = Counter((n.get("category"), n.get("direction")) for n in news if n.get("direction"))
    simple_dir = Counter(n.get("direction") for n in news if n.get("direction"))
    cities = Counter(n.get("city") for n in news if n.get("city")).most_common(6)
    times = sorted(n.get("time", "") for n in news if n.get("time"))
    span = [times[0][:10], times[-1][:10]] if times else ["", ""]

    # 每个 direction 归类到大类
    cat_for_dir = {}
    for (c, d) in dir_counter:
        cat_for_dir.setdefault(d, c)

    # 取出现频次≥2 的方向，最多 6 条作为趋势
    top = [d for d, _ in simple_dir.most_common() if simple_dir[d] >= 2][:6]
    trends = []
    for d in top:
        ins = DIRECTION_INSIGHTS.get(d, {"trend": "", "advice": ""})
        c = cat_for_dir.get(d, "")
        ev = [n.get("title") for n in news if n.get("direction") == d and n.get("title")][:2]
        trends.append({
            "dir": d, "cat": c, "count": simple_dir[d],
            "color": CAT_COLOR.get(c, "#26e0ff"),
            "trend": ins.get("trend", ""), "advice": ins.get("advice", ""), "evidence": ev,
        })

    # 角色建议：命中种子场景用种子文案；否则按关键词从趋势库动态生成（均在此固化，不在浏览器交互）
    if role in ROLE_ADVICE:
        rb = {"audience": role, "color": ROLE_ADVICE[role]["color"],
              "icon": ROLE_ADVICE[role]["icon"], "points": ROLE_ADVICE[role]["points"]}
    else:
        rb = role_advice_for(role, fallback_dirs=[t["dir"] for t in trends])

    return {
        "updated": span[1] or "", "total": total, "catCounts": dict(by_cat),
        "span": span, "topCities": cities, "trends": trends,
        "execAdvice": [rb],   # 单一角色块，已由 build 固化
    }

def norm_key(n):
    """去重主键：优先用链接（去掉 ?# 后的参数，转小写），否则用标题去空格。"""
    link = (n.get("link") or "").strip().lower()
    if link:
        return re.sub(r"[?#].*$", "", link)
    return re.sub(r"\s+", "", (n.get("title") or ""))

def dedup(news, keep_first=True):
    """按 norm_key 去重，保留首次出现的条目。"""
    seen, out = set(), []
    for n in news:
        k = norm_key(n)
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(n)
    return out

def load_news(source):
    if source == "supabase":
        import db
        print("从 Supabase 拉取…")
        return db.pull()
    if source == "sample":
        return json.loads(SAMPLE.read_text(encoding="utf-8"))
    # 默认：xlsx；不存在则回退 sample
    if XLSX.exists():
        return xlsx_to_news(XLSX)
    print("未找到 news.xlsx，使用 news_sample.json")
    return json.loads(SAMPLE.read_text(encoding="utf-8"))

def xlsx_to_news(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        d = dict(zip(header, r))
        try:
            coord = [float(d.get("lng")), float(d.get("lat"))]
        except (TypeError, ValueError):
            coord = [0.0, 0.0]
        out.append({
            "time": d.get("time"), "category": d.get("category"), "direction": d.get("direction"),
            "company": d.get("company"), "city": d.get("city"), "coord": coord,
            "title": d.get("title"), "link": d.get("link"),
            "event": d.get("event"), "why": d.get("why"),
            "innovation": d.get("innovation"), "learn": d.get("learn"),
            "is_sample": bool(d.get("is_sample")),
        })
    return out

def news_to_xlsx(news, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(COLUMNS)
    for n in news:
        row = []
        for c in COLUMNS:
            if c == "lng":
                row.append(n["coord"][0] if n.get("coord") else None)
            elif c == "lat":
                row.append(n["coord"][1] if n.get("coord") else None)
            elif c == "is_sample":
                row.append(bool(n.get("is_sample", False)))
            else:
                row.append(n.get(c))
        ws.append(row)
    wb.save(path)

def build_html(news, is_sample=False, source="xlsx", role=None):
    assert len(news) > 0, "无新闻数据"
    tpl = TEMPLATE.read_text(encoding="utf-8")
    echarts = (ASSETS / "echarts.min.js").read_text(encoding="utf-8")
    world = (ASSETS / "world.json").read_text(encoding="utf-8")
    china = (ASSETS / "china.json").read_text(encoding="utf-8")
    xlsx_lib = (ASSETS / "xlsx.full.min.js").read_text(encoding="utf-8")
    news_js = json.dumps(news, ensure_ascii=False, indent=2)
    sugg_js = json.dumps(build_suggestions(news, role=role), ensure_ascii=False, indent=2)
    out = (tpl
           .replace("__ECHARTS_LIB__", echarts)
           .replace("__WORLD_JSON__", world.strip())
           .replace("__CHINA_JSON__", china.strip())
           .replace("__XLSX_LIB__", xlsx_lib)
           .replace("__NEWS__", news_js)
           .replace("__SAMPLE__", "true" if is_sample else "false")
           .replace("__SUGGESTIONS__", sugg_js))
    for tok in ("__ECHARTS_LIB__","__WORLD_JSON__","__CHINA_JSON__","__XLSX_LIB__","__NEWS__","__SAMPLE__","__SUGGESTIONS__"):
        assert tok not in out, f"占位符未替换: {tok}"
    OUT_HTML.write_text(out, encoding="utf-8")
    print(f"已生成：{OUT_HTML}  ({OUT_HTML.stat().st_size//1024} KB, {len(news)} 条新闻)")

def build(source="xlsx", role=None):
    news = load_news(source)
    news = dedup(news)
    build_html(news, is_sample=(source == "sample"), source=source, role=role)

def cmd_init():
    news = json.loads(SAMPLE.read_text(encoding="utf-8"))
    for n in news:
        n["is_sample"] = True          # 标记为示例，replace 时只清这些行
    news_to_xlsx(news, XLSX)
    print(f"已生成可编辑数据源：{XLSX}  ({len(news)} 行，标记为示例数据)")

def cmd_pull():
    import db
    news = db.pull()
    news = dedup(news)
    news_to_xlsx(news, XLSX)
    print(f"已从 Supabase 拉取并写入：{XLSX}  ({len(news)} 行)")

def cmd_push():
    import db
    news = xlsx_to_news(XLSX)
    db.push(news)

def cmd_merge(json_path, clear=False):
    """并入真实新闻：clear=True 为 replace 模式，只移除标记为示例(is_sample)的行，
    绝不删除用户已整理的真实数据；均按内容去重。"""
    incoming = json.loads(pathlib.Path(json_path).read_text(encoding="utf-8"))
    existing = xlsx_to_news(XLSX) if XLSX.exists() else []
    if clear:
        total = len(existing)
        existing = [n for n in existing if not n.get("is_sample")]
        dropped = total - len(existing)
        print(f"replace 模式：已移除 {dropped} 条示例数据，保留 {len(existing)} 条真实数据")
    before = len(existing) + len(incoming)
    merged = dedup(existing + incoming)   # existing 在前，已存在条目不会被覆盖/重复
    skipped = before - len(merged)
    news_to_xlsx(merged, XLSX)
    build_html(merged, is_sample=False)
    print(f"已合并：新增 {len(incoming)} 条，跳过已存在 {skipped} 条，现有合计 {len(merged)} 条 -> 已重写 news.xlsx 与 HTML")

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)
    sub.add_parser("init")
    b = sub.add_parser("build"); b.add_argument("--source", default="xlsx", choices=["xlsx","supabase","sample"])
    b.add_argument("--role", default=None, help="建议区视角/角色，如 '浙江省分行行长'；不填用默认(金融高管视角)")
    sub.add_parser("pull")
    sub.add_parser("push")
    m = sub.add_parser("merge"); m.add_argument("json")
    r = sub.add_parser("replace"); r.add_argument("json")
    args = ap.parse_args()
    if args.cmd == "init": cmd_init()
    elif args.cmd == "build": build(args.source, args.role)
    elif args.cmd == "pull": cmd_pull()
    elif args.cmd == "push": cmd_push()
    elif args.cmd == "merge": cmd_merge(args.json, clear=False)
    elif args.cmd == "replace": cmd_merge(args.json, clear=True)

if __name__ == "__main__":
    main()
